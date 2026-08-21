import io
import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Turma, Student, StudentProgress
from app.ingestion.transform import calcular_alerta_aluno, TRILHAS

router = APIRouter(prefix="/api/v1/turma", tags=["relatorio-turma"])

NAVY = "#002364"
GREEN = "#10B981"

# Mesmo default usado em app/routers/dashboard.py — precisa bater com a
# mesma chave em app/ingestion/transform.py:TRILHAS.
TRILHA_PADRAO = "41"

STATUS_LABEL = {"concluido": "Concluído", "engajado": "Engajado", "inscrito": "Inscrito"}

# --- Ranking de Estudantes por performance ---
# Mesma fórmula do script de diagnóstico pré-semestre:
#   score = PESO_PROGRESSAO * progresso_pct
#         + PESO_PONTOS     * pontos_normalizados
#         + PESO_MOEDAS     * moedas_normalizados
# pontos/moedas são normalizados min-max pra escala 0-100 DENTRO DA
# TURMA antes de entrar na conta — sem isso, comparar pontos (que pode
# ir a milhares) direto com progresso (0-100%) não faz sentido.
PESO_PROGRESSAO = 0.60
PESO_PONTOS = 0.25
PESO_MOEDAS = 0.15


def _slug(texto: str) -> str:
    """Vira um nome de arquivo seguro (sem espaço/acento/caractere especial)."""
    limpo = re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_")
    return limpo or "turma"


def _normalizar_0_100(valores: list[float]) -> list[float]:
    """Normalização min-max pra escala 0-100. Turma inteira com o mesmo
    valor (ou lista vazia) -> todo mundo fica em 0 (sem variação pra medir)."""
    if not valores:
        return []
    minimo, maximo = min(valores), max(valores)
    if maximo == minimo:
        return [0.0] * len(valores)
    return [(v - minimo) / (maximo - minimo) * 100 for v in valores]


def _buscar_relatorio_turma(db: Session, nome: str, trilha: str = TRILHA_PADRAO) -> dict:
    turma = db.execute(select(Turma).where(Turma.name == nome)).scalar_one_or_none()
    if turma is None:
        raise HTTPException(status_code=404, detail=f"Turma '{nome}' não encontrada.")

    alunos = db.execute(select(Student).where(Student.turma_id == turma.id)).scalars().all()

    # Uma única query pra todos os alunos da turma, em vez de um SELECT de
    # StudentProgress por aluno dentro do loop (N+1 queries). Antes era
    # fixo em trail_external_id == "41" (Trilha Saldo+) — o relatório da
    # turma sempre mostrava o progresso da Saldo+ mesmo quando a Trilha
    # Pocket estava selecionada no dashboard. Agora respeita o mesmo
    # parâmetro `trilha` que /api/v1/dashboard já usa.
    progresso_por_aluno: dict[int, StudentProgress] = {}
    if alunos:
        progress_rows = db.execute(
            select(StudentProgress).where(
                StudentProgress.student_id.in_([a.id for a in alunos]),
                StudentProgress.trail_external_id == trilha,
            )
        ).scalars().all()
        progresso_por_aluno = {p.student_id: p for p in progress_rows}

    linhas = []
    for aluno in alunos:
        progresso = progresso_por_aluno.get(aluno.id)
        status = progresso.status if progresso else "inscrito"
        dias_sem_acesso, alerta, motivo_alerta = calcular_alerta_aluno(aluno.last_access)
        linhas.append({
            "nome": aluno.name or aluno.login,
            "login": aluno.login,
            "progresso_pct": round(progresso.progress_pct, 1) if progresso else 0.0,
            "pontos": aluno.pontos if aluno.pontos is not None else 0,
            "moedas": aluno.moedas if aluno.moedas is not None else 0,
            "status": STATUS_LABEL.get(status, status),
            "dias_sem_acesso": dias_sem_acesso,
            "alerta_inatividade": alerta,
            "motivo_alerta": motivo_alerta,
            "concluido_em": (
                progresso.completed_at.strftime("%d/%m/%Y")
                if progresso and progresso.completed_at else "—"
            ),
        })

    pontos_norm = _normalizar_0_100([l["pontos"] for l in linhas])
    moedas_norm = _normalizar_0_100([l["moedas"] for l in linhas])
    for linha, p_norm, m_norm in zip(linhas, pontos_norm, moedas_norm):
        linha["score_engajamento"] = round(
            PESO_PROGRESSAO * linha["progresso_pct"] + PESO_PONTOS * p_norm + PESO_MOEDAS * m_norm, 2
        )

    linhas.sort(key=lambda l: l["score_engajamento"], reverse=True)
    for posicao, linha in enumerate(linhas, start=1):
        linha["posicao_na_turma"] = posicao

    return {
        "turma": turma.name,
        "escola": turma.school.name if turma.school else turma.name,
        "trilha": TRILHAS.get(trilha, trilha),
        "total_alunos": len(linhas),
        "engajados": sum(1 for l in linhas if l["progresso_pct"] > 0),
        "concluintes": sum(1 for l in linhas if l["progresso_pct"] >= 100),
        "em_alerta": sum(1 for l in linhas if l["alerta_inatividade"]),
        "alunos": linhas,
    }


@router.get("/relatorio")
def get_relatorio_turma(nome: str, trilha: str = TRILHA_PADRAO, db: Session = Depends(get_db)):
    """Dados do relatório de uma turma — usado pelo modal 'Ver Alunos' no dashboard.
    Passe ?trilha=43 para o progresso na Trilha Pocket (default: 41, Saldo+)."""
    return _buscar_relatorio_turma(db, nome, trilha)


@router.get("/relatorio/pdf")
def get_relatorio_turma_pdf(nome: str, trilha: str = TRILHA_PADRAO, db: Session = Depends(get_db)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    dados = _buscar_relatorio_turma(db, nome, trilha)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("TituloSaldo", parent=styles["Title"], textColor=colors.HexColor(NAVY), fontSize=18)
    subtitulo_style = ParagraphStyle("SubtituloSaldo", parent=styles["Normal"], textColor=colors.HexColor("#64748B"), fontSize=9)

    story = [
        Paragraph(f"{dados['trilha']} — Relatório da turma {dados['turma']}", titulo_style),
        Paragraph(f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} (UTC)", subtitulo_style),
        Spacer(1, 14),
        Paragraph(
            f"{dados['total_alunos']} alunos · {dados['engajados']} engajados · {dados['concluintes']} concluintes",
            styles["Normal"],
        ),
        Spacer(1, 14),
    ]

    linhas_tabela = [["#", "Aluno", "Login", "Progresso", "Pontos", "Moedas", "Status", "Alerta", "Concluído em"]]
    for a in dados["alunos"]:
        linhas_tabela.append([
            str(a["posicao_na_turma"]), a["nome"], a["login"], f"{a['progresso_pct']:.1f}%",
            f"{a['pontos']:.0f}", f"{a['moedas']:.0f}", a["status"],
            a["motivo_alerta"] or "—", a["concluido_em"],
        ])

    tabela = Table(linhas_tabela, colWidths=[0.8 * cm, 3.6 * cm, 2.8 * cm, 1.9 * cm, 1.6 * cm, 1.6 * cm, 2 * cm, 3 * cm, 2.1 * cm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tabela)

    doc.build(story)
    buffer.seek(0)

    # Slug da trilha no nome do arquivo pra não sobrescrever o PDF de
    # Saldo+ com o de Pocket (ou vice-versa) da mesma turma.
    filename = f"relatorio_{_slug(dados['turma'])}_{_slug(dados['trilha'])}.pdf"
    return StreamingResponse(
        buffer, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/relatorio/excel")
def get_relatorio_turma_excel(nome: str, trilha: str = TRILHA_PADRAO, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    dados = _buscar_relatorio_turma(db, nome, trilha)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório da Turma"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{dados['trilha']} — Relatório da turma {dados['turma']}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws["A2"] = (
        f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} (UTC) — "
        f"{dados['total_alunos']} alunos · {dados['engajados']} engajados · {dados['concluintes']} concluintes"
    )
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    header_row = 4
    for col, texto in enumerate(["#", "Aluno", "Login", "Progresso", "Pontos", "Moedas", "Status", "Alerta", "Concluído em"], start=1):
        celula = ws.cell(row=header_row, column=col, value=texto)
        celula.font = Font(name="Arial", bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
        celula.alignment = Alignment(horizontal="center")

    for i, aluno in enumerate(dados["alunos"], start=header_row + 1):
        ws.cell(row=i, column=1, value=aluno["posicao_na_turma"]).font = Font(name="Arial")
        ws.cell(row=i, column=2, value=aluno["nome"]).font = Font(name="Arial")
        ws.cell(row=i, column=3, value=aluno["login"]).font = Font(name="Arial")
        celula_pct = ws.cell(row=i, column=4, value=aluno["progresso_pct"] / 100)
        celula_pct.number_format = "0.0%"
        celula_pct.font = Font(name="Arial")
        ws.cell(row=i, column=5, value=aluno["pontos"]).font = Font(name="Arial")
        ws.cell(row=i, column=6, value=aluno["moedas"]).font = Font(name="Arial")
        ws.cell(row=i, column=7, value=aluno["status"]).font = Font(name="Arial")
        celula_alerta = ws.cell(row=i, column=8, value=aluno["motivo_alerta"] or "—")
        celula_alerta.font = Font(name="Arial", color="DC2626" if aluno["alerta_inatividade"] else "000000")
        ws.cell(row=i, column=9, value=aluno["concluido_em"]).font = Font(name="Arial")

    for col, largura in zip("ABCDEFGHI", [5, 28, 20, 12, 10, 10, 14, 20, 16]):
        ws.column_dimensions[col].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_{_slug(dados['turma'])}_{_slug(dados['trilha'])}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )